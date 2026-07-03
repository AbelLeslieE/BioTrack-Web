# ==========================================
# IMPORTS
# ==========================================

from openpyxl import load_workbook

from backend.models.inventory import Inventory


# ==========================================
# IMPORT INVENTORY EXCEL
# ==========================================

def import_inventory_excel(file, db):

    workbook = load_workbook(file, data_only=True)

    added = 0
    updated = 0
    skipped = 0

    # Skip non-stock sheets
    ignore_sheets = [
        "INDEX",
        "NEW STOCK INVOICE DETAIS"
    ]

    for sheet in workbook.worksheets:

        if sheet.title.strip().upper() in [x.upper() for x in ignore_sheets]:
            continue

        bpk_number = sheet.title.strip()

        item_description = sheet["A1"].value or bpk_number

        item_code = sheet["C3"].value if sheet["C3"].value else ""

        current_stock = 0

        # Closing Stock is Column H
        for row in range(sheet.max_row, 2, -1):

            value = sheet[f"H{row}"].value

            if value is not None:

                try:

                    current_stock = int(value)

                except:

                    current_stock = 0

                break

        item = (

            db.query(Inventory)

            .filter(

                Inventory.bpk_number == bpk_number

            )

            .first()

        )

        if item:

            item.item_name = item_description

            item.current_stock = current_stock

            updated += 1

        else:

            db.add(

                Inventory(

                    bpk_number=bpk_number,

                    item_name=item_description,

                    current_stock=current_stock,

                    minimum_stock=5,

                    maximum_stock=100,

                    category="Biomedical Spare",

                    manufacturer="",

                    supplier="",

                    location="Main Store",

                    rack="",

                    purchase_price=0,

                    remarks="Imported from Stock Workbook"

                )

            )

            added += 1

    db.commit()

    return {

        "success": True,

        "added": added,

        "updated": updated,

        "skipped": skipped

    }